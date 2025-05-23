from utils.logger import app_logger
import pandas as pd
from utils.logger import app_logger
from utils.config_loader import get_template_config
from logic.manage.utils.csv_loader import load_all_filtered_dataframes
from logic.manage.utils.load_template import load_master_and_template
from config.loader.main_path import MainPath
from logic.manage.readers.read_transport_discount import ReadTransportDiscount
import streamlit as st
import time
import re


from logic.manage.processors.block_unit_price.process0 import (
    make_df_shipping_after_use,
    apply_unit_price_addition,
    apply_transport_fee_by1,
)
from logic.manage.processors.block_unit_price.process1 import (
    create_transport_selection_form,
)
from logic.manage.processors.block_unit_price.process2 import (
    confirm_transport_selection,
)

# デバッグ用
from utils.debug_tools import debug_pause


def process(dfs):
    """
    ブロック単価計算の主処理を行う関数

    Args:
        dfs: 入力データフレーム群

    Returns:
        pd.DataFrame: 計算結果を含むマスターCSV
    """
    logger = app_logger()

    # --- 内部ステップ管理 ---
    mini_step = st.session_state.get("process_mini_step", 0)

    # --- 設定とマスターデータの読み込み ---
    # テンプレート設定
    template_key = "block_unit_price"
    template_config = get_template_config()[template_key]
    template_name = template_config["key"]
    csv_keys = template_config["required_files"]
    logger.info(f"[テンプレート設定読込] key={template_key}, files={csv_keys}")

    # マスターデータ
    config = get_template_config()["block_unit_price"]
    master_path = config["master_csv_path"]["vendor_code"]
    master_csv = load_master_and_template(master_path)

    # 運搬費データ
    mainpath = MainPath()
    reader = ReadTransportDiscount(mainpath)
    df_transport = reader.load_discounted_df()

    # 出荷データ
    df_dict = load_all_filtered_dataframes(dfs, csv_keys, template_name)
    df_shipping = df_dict.get("shipping")

    # --- 段階的処理の実行 ---
    if mini_step == 0:
        _process_step0(df_shipping, master_csv, df_transport)
        return None
    elif mini_step == 1:
        _process_step1(df_transport)
        return None
    elif mini_step == 2:
        return _process_step2(df_shipping, df_transport)

    return None


def _process_step0(
    df_shipping: pd.DataFrame, master_csv: pd.DataFrame, df_transport: pd.DataFrame
) -> None:
    """繰り返し不必要な処理
        基本データ処理と運搬費（固定のもの、社数１）を行うステップ0の実装
        繰り返す必要がないため、このステップに入れる。

    Args:
        df_shipping: 出荷データ
        master_csv: マスターデータ
        df_transport: 運搬費データ
    """
    logger = app_logger()
    logger.info("▶️ Step0: フィルタリング・単価追加・固定運搬費")

    df_shipping = make_df_shipping_after_use(master_csv, df_shipping)  # フィルタリング
    df_shipping = apply_unit_price_addition(master_csv, df_shipping)  # 単価追加
    df_shipping = apply_transport_fee_by1(df_shipping, df_transport)  # 固定運搬費

    st.session_state.df_shipping_first = df_shipping
    st.session_state.process_mini_step = 1
    st.rerun()


def _process_step1(df_transport: pd.DataFrame) -> None:
    """繰り返し必要な処理
        運搬業者選択を行うステップ1の実装

    処理の流れ:
        1. 前のステップで保存した出荷データを取得
        2. 運搬業者の選択状態を確認
            - 未選択：選択フォームを表示して選択を促す
            - 選択済：次のステップに進む
        3. 選択結果をセッションに保存してページを再読み込み

    Args:
        df_transport: 運搬費データ
    """
    # ロガーの初期化
    logger = app_logger()
    logger.info("▶️ Step1: 選択式運搬費")

    # ステップ1: 前のステップの出荷データを取得
    df_after = st.session_state.df_shipping_first

    # ステップ2: 運搬業者の選択状態を確認
    if not st.session_state.get("block_unit_price_confirmed", False):
        # 未選択の場合：選択フォームを表示
        df_after = create_transport_selection_form(df_after, df_transport)

        # ステップ3: 選択結果を保存して再読み込み
        st.session_state.df_shipping = df_after
        st.rerun()
    else:
        # 選択済みの場合：次のステップへ進む
        logger.info("▶️ 選択済みなのでスキップ")
        st.session_state.process_mini_step = 2
        st.rerun()


def _process_step2(
    df_shipping: pd.DataFrame, df_transport: pd.DataFrame
) -> pd.DataFrame:
    """最終計算処理を行うステップ2の実装

    Args:
        df_shipping: 出荷データ
        df_transport: 運搬費データ

    Returns:
        pd.DataFrame: 計算結果を含むマスターCSV
    """
    logger = app_logger()
    logger.info("▶️ Step2: 加算処理実行中")
    df_after = st.session_state.df_shipping

    # 運搬業者選択の確認
    confirm_transport_selection(df_after)

    # 運搬費の計算と適用
    df_after = apply_transport_fee_by_vendor(df_after, df_transport)
    df_after = apply_weight_based_transport_fee(df_after, df_transport)

    # ブロック単価の計算と表示用データの作成
    df_after = make_total_sum(df_after)
    df_after = df_cul_filtering(df_after)
    master_csv = first_cell_in_template(df_after)
    master_csv = make_sum_date(master_csv, df_shipping)

    # ステートの初期化
    st.session_state.process_mini_step = 0

    return master_csv


def apply_transport_fee_by_vendor(
    df_after: pd.DataFrame, df_transport: pd.DataFrame
) -> pd.DataFrame:
    """運搬業者ごとの運搬費を適用する関数

    Args:
        df_after: 処理対象の出荷データフレーム
        df_transport: 運搬費データフレーム

    Returns:
        pd.DataFrame: 運搬費が適用された出荷データフレーム
    """
    from logic.manage.utils.column_utils import apply_column_addition_by_keys

    # 運搬業者が設定されている行を抽出
    target_rows = df_after[df_after["運搬業者"].notna()].copy()

    # 運搬費の適用（業者CDで結合）
    updated_target_rows = apply_column_addition_by_keys(
        base_df=target_rows,
        addition_df=df_transport,
        join_keys=["業者CD", "運搬業者"],
        value_col_to_add="運搬費",
        update_target_col="運搬費",
    )

    # 運搬業者が未設定の行を保持
    non_transport_rows = df_after[df_after["運搬業者"].isna()].copy()

    # 処理済みデータの結合
    df_after = pd.concat([updated_target_rows, non_transport_rows], ignore_index=True)

    return df_after


def apply_weight_based_transport_fee(
    df_after: pd.DataFrame, df_transport: pd.DataFrame
) -> pd.DataFrame:
    """重量に基づく運搬費を計算して適用する関数

    Args:
        df_after: 処理対象の出荷データフレーム
        df_transport: 運搬費データフレーム（"数字*weight"形式の運搬費を含む）

    Returns:
        pd.DataFrame: 重量に基づく運搬費が適用された出荷データフレーム
    """
    # 重量ベースの運搬費行を抽出
    transport_fee_col = (
        df_transport["運搬費"].astype(str).str.replace(r"\s+", "", regex=True)
    )
    weight_based_mask = transport_fee_col.str.fullmatch(r"\d+\*weight", na=False)
    weight_based_transport = df_transport[weight_based_mask].copy()

    # 運搬費係数の抽出と変換
    weight_based_transport["運搬費係数"] = (
        weight_based_transport["運搬費"].str.extract(r"^(\d+)")[0].astype(float)
    )

    # 必要な列の選択と重複除去
    weight_based_transport = weight_based_transport.drop_duplicates(
        subset=["業者CD", "運搬業者"]
    )[["業者CD", "運搬業者", "運搬費係数"]]

    # 運搬費係数の適用
    df_result = df_after.merge(
        weight_based_transport,
        how="left",
        on=["業者CD", "運搬業者"],
        suffixes=("", "_formula"),
    )

    # 重量ベースの運搬費計算
    has_coefficient_mask = df_result["運搬費係数"].notna()
    df_result.loc[has_coefficient_mask, "運搬費"] = (
        df_result.loc[has_coefficient_mask, "運搬費係数"]
        * df_result.loc[has_coefficient_mask, "正味重量"]
    ).astype(float)

    return df_result


def make_total_sum(df):

    # 総額
    df["総額"] = df["単価"] * df["正味重量"] + df["運搬費"]
    df["ブロック単価"] = (df["総額"] / df["正味重量"].replace(0, pd.NA)).round(2)
    return df


def df_cul_filtering(df):
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

    # dfカラムのフィルタリング
    df = df[["業者名", "明細備考", "正味重量", "総額", "ブロック単価"]]

    #     # カラム名の変更
    #     df = df.rename(columns={
    #     # "業者名": "取引先名",
    #     "明細備考": "明細備考",
    #     "正味重量": "数量",
    #     "総額": "金額",
    #     "ブロック単価": "単価"
    # })
    return df


def first_cell_in_template(df):

    start_row = 7
    full_col_to_cell = {
        "業者名": "B",
        "明細備考": "C",
        "正味重量": "D",
        "総額": "E",
        "ブロック単価": "F",
    }

    # セル情報を再構築
    full_cell_info = []

    for idx, row in df.iterrows():
        for col, col_letter in full_col_to_cell.items():
            cell = f"{col_letter}{start_row + idx}"
            value = row[col]
            full_cell_info.append({"大項目": col, "セル": cell, "値": value})

    full_cell_df = pd.DataFrame(full_cell_info)

    return full_cell_df


def make_sum_date(df, df_shipping):
    from utils.date_tools import to_reiwa_format

    # 日付を令和表記に変換（例: "令和6年5月16日"）
    date = to_reiwa_format(df_shipping["伝票日付"].iloc[0])

    # 追加行を定義
    new_row = pd.DataFrame([{"大項目": "日付", "セル": "E4", "値": date}])

    # df に行を追加
    df = pd.concat([df, new_row], ignore_index=True)

    return df


def calculate_block_unit_price(df: pd.DataFrame) -> pd.DataFrame:
    """ブロック単価を計算する関数

    Args:
        df: 処理対象のデータフレーム

    Returns:
        pd.DataFrame: ブロック単価が計算されたデータフレーム
    """
    # 総額の計算（単価 × 正味重量 + 運搬費）
    df["総額"] = df["単価"] * df["正味重量"] + df["運搬費"]

    # ブロック単価の計算（総額 ÷ 正味重量）、0除算を回避
    df["ブロック単価"] = (df["総額"] / df["正味重量"].replace(0, pd.NA)).round(2)
    return df


def filter_display_columns(df: pd.DataFrame) -> pd.DataFrame:
    """表示用の列を選択する関数

    Args:
        df: 処理対象のデータフレーム

    Returns:
        pd.DataFrame: 表示用に列が選択されたデータフレーム
    """
    display_columns = ["業者名", "明細備考", "正味重量", "総額", "ブロック単価"]
    return df[display_columns]


def create_cell_mapping(df: pd.DataFrame) -> pd.DataFrame:
    """データフレームの値をExcelセルにマッピングする関数

    Args:
        df: 処理対象のデータフレーム

    Returns:
        pd.DataFrame: セルマッピング情報を含むデータフレーム
    """
    start_row = 7
    column_to_cell = {
        "業者名": "B",
        "明細備考": "C",
        "正味重量": "D",
        "総額": "E",
        "ブロック単価": "F",
    }

    # セルマッピング情報の作成
    cell_mappings = []
    for idx, row in df.iterrows():
        for column, cell_letter in column_to_cell.items():
            cell_position = f"{cell_letter}{start_row + idx}"
            cell_mappings.append(
                {"大項目": column, "セル": cell_position, "値": row[column]}
            )

    return pd.DataFrame(cell_mappings)


def add_date_information(df: pd.DataFrame, df_shipping: pd.DataFrame) -> pd.DataFrame:
    """日付情報を追加する関数

    Args:
        df: セルマッピング情報を含むデータフレーム
        df_shipping: 出荷データフレーム

    Returns:
        pd.DataFrame: 日付情報が追加されたデータフレーム
    """
    from utils.date_tools import to_reiwa_format

    # 伝票日付を令和形式に変換
    reiwa_date = to_reiwa_format(df_shipping["伝票日付"].iloc[0])

    # 日付情報の追加
    date_row = pd.DataFrame([{"大項目": "日付", "セル": "E4", "値": reiwa_date}])

    return pd.concat([df, date_row], ignore_index=True)
