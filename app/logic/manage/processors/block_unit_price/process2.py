import streamlit as st
import pandas as pd


def confirm_transport_selection(df_after: pd.DataFrame) -> None:
    """運搬業者の選択内容を確認するダイアログを表示する

    処理の流れ:
        1. 選択された運搬業者の一覧を表示
        2. 確認用のYes/Noボタンを表示
        3. ユーザーの選択に応じて処理を分岐
            - Yes: 次のステップへ進む（process_mini_step = 2）
            - No: Step1（選択画面）に戻る（process_mini_step = 1）

    Args:
        df_after (pd.DataFrame): 運搬業者が選択された出荷データ
    """
    # セッション状態の初期化
    if "transport_selection_confirmed" not in st.session_state:
        st.session_state.transport_selection_confirmed = False

    def _create_confirmation_view(df: pd.DataFrame) -> pd.DataFrame:
        """確認用の表示データを作成"""
        filtered_df = df[df["運搬業者"].notna()]
        return filtered_df[["業者名", "品名", "明細備考", "運搬業者"]]

    def _show_confirmation_buttons() -> tuple[bool, bool]:
        """確認用のYes/Noボタンを表示"""
        st.write("この運搬業者選択で確定しますか？")
        col1, col2 = st.columns([1, 1])

        with col1:
            yes_clicked = st.button("✅ はい（この内容で確定）", key="yes_button")
        with col2:
            no_clicked = st.button("🔁 いいえ（やり直す）", key="no_button")

        return yes_clicked, no_clicked

    def _handle_user_selection(yes_clicked: bool, no_clicked: bool) -> None:
        """ユーザーの選択結果を処理"""
        if yes_clicked:
            st.success("✅ 確定されました。次に進みます。")
            st.session_state.transport_selection_confirmed = True
            st.session_state.process_mini_step = 2
            st.rerun()

        if no_clicked:
            st.warning("🔁 選択をやり直します（Step1に戻ります）")
            st.session_state.transport_selection_confirmed = False
            st.session_state.process_mini_step = 1
            st.rerun()

    # すでに確認済みの場合はスキップ
    if st.session_state.transport_selection_confirmed:
        return

    # メイン処理の実行
    st.title("運搬業者の確認")

    # 1. 確認用データの表示
    df_view = _create_confirmation_view(df_after)
    st.dataframe(df_view)

    # 2. 確認ボタンの表示と選択結果の取得
    yes_clicked, no_clicked = _show_confirmation_buttons()

    # 3. 選択結果の処理
    _handle_user_selection(yes_clicked, no_clicked)

    # 4. ユーザーの操作待ち
    st.stop()


def apply_transport_fee_by_vendor(
    df_after: pd.DataFrame, df_transport: pd.DataFrame
) -> pd.DataFrame:
    """運搬業者ごとの運搬費を適用する関数

    Args:
        df_after: 処理対象の出荷データフレーム
        df_transport: 運搬費データフレーム

    Returns:
        pd.DataFrame: 運搬費が適用された出荷データフレーム
    """
    from logic.manage.utils.column_utils import apply_column_addition_by_keys

    # 運搬業者が設定されている行を抽出
    target_rows = df_after[df_after["運搬業者"].notna()].copy()

    # 運搬費の適用（業者CDで結合）
    updated_target_rows = apply_column_addition_by_keys(
        base_df=target_rows,
        addition_df=df_transport,
        join_keys=["業者CD", "運搬業者"],
        value_col_to_add="運搬費",
        update_target_col="運搬費",
    )

    # 運搬業者が未設定の行を保持
    non_transport_rows = df_after[df_after["運搬業者"].isna()].copy()

    # 処理済みデータの結合
    df_after = pd.concat([updated_target_rows, non_transport_rows], ignore_index=True)

    return df_after


def apply_weight_based_transport_fee(
    df_after: pd.DataFrame, df_transport: pd.DataFrame
) -> pd.DataFrame:
    """重量に基づく運搬費を計算して適用する関数

    Args:
        df_after: 処理対象の出荷データフレーム
        df_transport: 運搬費データフレーム（"数字*weight"形式の運搬費を含む）

    Returns:
        pd.DataFrame: 重量に基づく運搬費が適用された出荷データフレーム
    """
    # 重量ベースの運搬費行を抽出
    transport_fee_col = (
        df_transport["運搬費"].astype(str).str.replace(r"\s+", "", regex=True)
    )
    weight_based_mask = transport_fee_col.str.fullmatch(r"\d+\*weight", na=False)
    weight_based_transport = df_transport[weight_based_mask].copy()

    # 運搬費係数の抽出と変換
    weight_based_transport["運搬費係数"] = (
        weight_based_transport["運搬費"].str.extract(r"^(\d+)")[0].astype(float)
    )

    # 必要な列の選択と重複除去
    weight_based_transport = weight_based_transport.drop_duplicates(
        subset=["業者CD", "運搬業者"]
    )[["業者CD", "運搬業者", "運搬費係数"]]

    # 運搬費係数の適用
    df_result = df_after.merge(
        weight_based_transport,
        how="left",
        on=["業者CD", "運搬業者"],
        suffixes=("", "_formula"),
    )

    # 重量ベースの運搬費計算
    has_coefficient_mask = df_result["運搬費係数"].notna()
    df_result.loc[has_coefficient_mask, "運搬費"] = (
        df_result.loc[has_coefficient_mask, "運搬費係数"]
        * df_result.loc[has_coefficient_mask, "正味重量"]
    ).astype(float)

    return df_result


def make_total_sum(df, master_csv):
    # 個々の金額計算と計算用重量の設定
    def calculate_row(row):
        if row["単位名"] == "kg":
            row["金額"] = row["単価"] * row["正味重量"]
        elif row["単位名"] == "台":
            row["金額"] = row["単価"] * row["数量"]
        return row

    # 行ごとに計算を適用
    df = df.apply(calculate_row, axis=1)

    # 総額の計算
    df["総額"] = df["金額"] + df["運搬費"]

    # ブロック単価の計算（計算用重量を使用）
    df["ブロック単価"] = (df["総額"] / df["正味重量"].replace(0, pd.NA)).round(2)

    return df


def df_cul_filtering(df):
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

    # dfカラムのフィルタリング
    df = df[["業者名", "明細備考", "正味重量", "総額", "ブロック単価"]]

    #     # カラム名の変更
    #     df = df.rename(columns={
    #     # "業者名": "取引先名",
    #     "明細備考": "明細備考",
    #     "正味重量": "数量",
    #     "総額": "金額",
    #     "ブロック単価": "単価"
    # })
    return df


def first_cell_in_template(df):

    start_row = 7
    full_col_to_cell = {
        "業者名": "B",
        "明細備考": "C",
        "正味重量": "D",
        "総額": "E",
        "ブロック単価": "F",
    }

    # セル情報を再構築
    full_cell_info = []

    for idx, row in df.iterrows():
        for col, col_letter in full_col_to_cell.items():
            cell = f"{col_letter}{start_row + idx}"
            value = row[col]
            full_cell_info.append({"大項目": col, "セル": cell, "値": value})

    full_cell_df = pd.DataFrame(full_cell_info)

    return full_cell_df


def make_sum_date(df, df_shipping):
    from utils.date_tools import to_reiwa_format

    # 日付を令和表記に変換（例: "令和6年5月16日"）
    date = to_reiwa_format(df_shipping["伝票日付"].iloc[0])

    # 追加行を定義
    new_row = pd.DataFrame([{"大項目": "日付", "セル": "E4", "値": date}])

    # df に行を追加
    df = pd.concat([df, new_row], ignore_index=True)

    return df


def calculate_block_unit_price(df: pd.DataFrame) -> pd.DataFrame:
    """ブロック単価を計算する関数

    Args:
        df: 処理対象のデータフレーム

    Returns:
        pd.DataFrame: ブロック単価が計算されたデータフレーム
    """
    # 総額の計算（単価 × 正味重量 + 運搬費）
    df["総額"] = df["単価"] * df["正味重量"] + df["運搬費"]

    # ブロック単価の計算（総額 ÷ 正味重量）、0除算を回避
    df["ブロック単価"] = (df["総額"] / df["正味重量"].replace(0, pd.NA)).round(2)
    return df


def filter_display_columns(df: pd.DataFrame) -> pd.DataFrame:
    """表示用の列を選択する関数

    Args:
        df: 処理対象のデータフレーム

    Returns:
        pd.DataFrame: 表示用に列が選択されたデータフレーム
    """
    display_columns = ["業者名", "明細備考", "正味重量", "総額", "ブロック単価"]
    return df[display_columns]


def create_cell_mapping(df: pd.DataFrame) -> pd.DataFrame:
    """データフレームの値をExcelセルにマッピングする関数

    Args:
        df: 処理対象のデータフレーム

    Returns:
        pd.DataFrame: セルマッピング情報を含むデータフレーム
    """
    start_row = 7
    column_to_cell = {
        "業者名": "B",
        "明細備考": "C",
        "正味重量": "D",
        "総額": "E",
        "ブロック単価": "F",
    }

    # セルマッピング情報の作成
    cell_mappings = []
    for idx, row in df.iterrows():
        for column, cell_letter in column_to_cell.items():
            cell_position = f"{cell_letter}{start_row + idx}"
            cell_mappings.append(
                {"大項目": column, "セル": cell_position, "値": row[column]}
            )

    return pd.DataFrame(cell_mappings)


def add_date_information(df: pd.DataFrame, df_shipping: pd.DataFrame) -> pd.DataFrame:
    """日付情報を追加する関数

    Args:
        df: セルマッピング情報を含むデータフレーム
        df_shipping: 出荷データフレーム

    Returns:
        pd.DataFrame: 日付情報が追加されたデータフレーム
    """
    from utils.date_tools import to_reiwa_format

    # 伝票日付を令和形式に変換
    reiwa_date = to_reiwa_format(df_shipping["伝票日付"].iloc[0])

    # 日付情報の追加
    date_row = pd.DataFrame([{"大項目": "日付", "セル": "E4", "値": reiwa_date}])

    return pd.concat([df, date_row], ignore_index=True)
