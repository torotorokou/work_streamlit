from openpyxl import load_workbook
from io import BytesIO
import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl.cell.cell import MergedCell
from utils.logger import app_logger
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from copy import copy  # ← 追加

def safe_excel_value(value):
    """Excelに書き込める形式に変換するユーティリティ関数"""
    if pd.isna(value) or value is pd.NA or value is np.nan:
        return None
    elif isinstance(value, (dict, list, set)):
        return str(value)
    elif hasattr(value, "strftime"):
        return value.strftime("%Y/%m/%d")
    return value


def load_template_workbook(template_path: str) -> Workbook:
    return load_workbook(template_path)


def write_dataframe_to_worksheet(df: pd.DataFrame, ws: Worksheet, logger=None):
    if logger is None:
        logger = app_logger()

    for idx, row in df.iterrows():
        cell_ref = row.get("セル")
        value = safe_excel_value(row.get("値"))

        if pd.isna(cell_ref) or str(cell_ref).strip() in ["", "未設定"]:
            logger.info(f"空欄または未設定のセルはスキップされました。行 {idx}")
            continue

        try:
            cell = ws[cell_ref]

            if isinstance(cell, MergedCell):
                logger.warning(f"セル {cell_ref} は結合セルで書き込み不可。値: {value}")
                continue

            # --- 書式をdeep copyで保持 ---
            original_font = copy(cell.font)
            original_fill = copy(cell.fill)
            original_border = copy(cell.border)
            original_format = cell.number_format

            # 値の上書き
            cell.value = value

            # --- 書式の復元 ---
            cell.font = original_font
            cell.fill = original_fill
            cell.border = original_border
            cell.number_format = original_format

        except Exception as e:
            logger.error(f"セル {cell_ref} 書き込み失敗: {e} / 値: {value}")


def rename_sheet(wb: Workbook, new_title: str):
    ws = wb.active
    ws.title = new_title


def save_workbook_to_bytesio(wb: Workbook) -> BytesIO:
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def write_values_to_template(
    df: pd.DataFrame, template_path: str, extracted_date: str
) -> BytesIO:
    """
    単一責任原則に基づいて分割されたExcelテンプレート書き込み関数
    - テンプレ読み込み
    - セルへの書き込み
    - シート名変更
    - メモリ出力
    """
    logger = app_logger()
    wb = load_template_workbook(template_path)
    ws = wb.active

    write_dataframe_to_worksheet(df, ws, logger=logger)
    rename_sheet(wb, extracted_date)
    return save_workbook_to_bytesio(wb)
